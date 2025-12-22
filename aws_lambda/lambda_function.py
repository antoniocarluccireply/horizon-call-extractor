import os
import re
import base64
import mimetypes
import uuid
import json
import calendar
from typing import Dict, List, Optional
import boto3
import urllib.request
import urllib.error
import traceback
from datetime import date
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Alignment

from parser_horizon import parse_calls
from parser_edf import parse_edf

s3 = boto3.client(
    "s3",
    region_name="eu-central-1",
    endpoint_url="https://s3.eu-central-1.amazonaws.com",
        )

BUCKET = os.environ.get("BUCKET_NAME", "")
def _require_bucket():
    if not BUCKET:
        # Non blocchiamo la UI, ma blocchiamo le API che richiedono S3
        raise RuntimeError("Missing env var BUCKET_NAME")

# --- OpenAI (optional) ---
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")  # keep secret in Lambda env vars
OPENAI_MODEL = os.environ.get("OPENAI_MODEL", "gpt-5-mini")
OPENAI_MAX_TOPICS = int(os.environ.get("OPENAI_MAX_TOPICS", "25"))
OPENAI_BODY_MAX_CHARS = int(os.environ.get("OPENAI_BODY_MAX_CHARS", "6000"))
DEFAULT_MIN_BUDGET_M = float(os.environ.get("DEFAULT_MIN_BUDGET_M", "0"))
DOC_HORIZON = "horizon"
DOC_EDF = "edf"
UI_PATH = os.path.join(os.path.dirname(__file__), "ui.html")
ASSETS_DIR = os.path.join(os.path.dirname(__file__), "assets")

def _serve_asset(request_path: str):
    """
    Serve files packaged inside Lambda under aws_lambda/assets/*
    Example: GET /assets/adeptic.png
    """
    # request_path comes like "/assets/adeptic.png"
    rel = (request_path or "").lstrip("/")  # "assets/adeptic.png"
    if not rel.startswith("assets/"):
        return {"statusCode": 404, "headers": {"Content-Type": "text/plain"}, "body": "Not found"}

    # Avoid path traversal
    rel_file = rel[len("assets/"):]  # "adeptic.png"
    abs_path = os.path.abspath(os.path.join(ASSETS_DIR, rel_file))
    assets_root = os.path.abspath(ASSETS_DIR)

    if not abs_path.startswith(assets_root + os.sep):
        return {"statusCode": 403, "headers": {"Content-Type": "text/plain"}, "body": "Forbidden"}

    if not os.path.exists(abs_path):
        return {"statusCode": 404, "headers": {"Content-Type": "text/plain"}, "body": "Not found"}

    with open(abs_path, "rb") as f:
        data = f.read()

    ctype, _ = mimetypes.guess_type(abs_path)
    return {
        "statusCode": 200,
        "headers": {
            "Content-Type": ctype or "application/octet-stream",
            "Cache-Control": "public, max-age=86400",
            "access-control-allow-origin": "*",
        },
        "isBase64Encoded": True,
        "body": base64.b64encode(data).decode("utf-8"),
    }
EDF_CALL_FAMILY_LABELS = {
    "RA": "RA — Research Actions",
    "DA": "DA — Development Actions",
    "CSA": "CSA — Coordination & Support Actions",
}

try:
    with open(UI_PATH, "r", encoding="utf-8") as f:
        HTML = f.read()
except FileNotFoundError:
    HTML = "<html><body><p>UI not found</p></body></html>"


def _safe_base_name(file_name: str) -> str:
    base = os.path.basename(file_name or "").strip()
    if not base:
        return "file"

    base = re.sub(r"[\\/:*?\"<>|]", "_", base)
    base = base.strip(". ")
    name, _ext = os.path.splitext(base)
    cleaned = name or "file"
    return cleaned[:120]


def _topic_url(topic_id: str) -> str:
    tid = (topic_id or "").strip()
    return (
        f"https://ec.europa.eu/info/funding-tenders/opportunities/portal/screen/opportunities/topic-details/{tid}"
        if tid
        else ""
        )


def _edf_call_family_from_id(call_id: Optional[str]) -> Optional[str]:
    parts = (call_id or "").split("-")
    if len(parts) < 3:
        return None
    fam = parts[2].upper()
    return fam if fam in EDF_CALL_FAMILY_LABELS else None


def _edf_call_family_label(family: Optional[str]) -> Optional[str]:
    if not family:
        return None
    return EDF_CALL_FAMILY_LABELS.get(str(family).upper())


def _edf_scale_label(is_large_scale: bool) -> str:
    return "Large-scale" if is_large_scale else "Standard"


def _edf_is_large_scale(row: Dict) -> bool:
    def _has_ls(identifier: Optional[str]) -> bool:
        if not identifier:
            return False
        parts = identifier.split("-")
        return len(parts) >= 4 and any(p.upper() == "LS" for p in parts[3:])

    if _has_ls(row.get("topic_id")) or _has_ls(row.get("call_id")):
        return True

    text_blob = " ".join(
        [
            str(row.get("topic_title") or ""),
            str(row.get("topic_description_verbatim") or ""),
        ]
    )
    return bool(re.search(r"\blarge[-\s]?scale\b", text_blob, flags=re.IGNORECASE))


def _funding_percentage(row: Dict, doc_type: str) -> Optional[str]:
    """
    Return funding percentage as a string with '%' or None when not available.
    Rules from the functional spec:
    - Horizon: map from action_type; IA defaults to 70% unless "non-profit" is explicitly found.
    - EDF: only surface percentages explicitly present in the text (parser may provide numeric value).
    """
    if doc_type == DOC_EDF:
        val = row.get("funding_percentage")
        if isinstance(val, (int, float)) and val >= 0:
            return f"{val:g}%"
        if isinstance(val, str) and val.strip():
            return val.strip()
        return None

    action = (row.get("action_type") or "").strip().upper()
    if not action:
        return None

    if action in {"RIA", "CSA"}:
        return "100%"
    if action == "IA":
        body_text = " ".join(
            [
                str(row.get("topic_body") or ""),
                str(row.get("topic_description") or ""),
                str(row.get("topic_description_verbatim") or ""),
            ]
        ).lower()
        if "non-profit" in body_text or "non profit" in body_text:
            return "100%"
        return "70%"

    # PCP / PPI depend on the call text; we only surface if a value is explicitly available
    explicit = row.get("funding_percentage")
    if isinstance(explicit, (int, float)):
        return f"{explicit:g}%"
    if isinstance(explicit, str) and explicit.strip():
        return explicit.strip()
    return None


def _row_call_type(row: Dict, doc_type: str) -> Optional[str]:
    if doc_type == DOC_EDF:
        return row.get("call_type") or row.get("type_of_action") or row.get("call_family_display")
    return row.get("action_type")


def _row_min_budget(row: Dict) -> Optional[float]:
    for key in (
        "budget_per_project_min_eur_m",
        "budget_per_project_m",
        "indicative_budget_eur_m",
    ):
        v = row.get(key)
        if isinstance(v, (int, float)):
            return float(v)
    return None


def _process_pdf_keys(
    pdf_keys: List[str],
    context=None,
    call_types=None,
    min_budget_m=None,
    opening_filter: str = "",
    deadline_filter: str = "",
    original_names: Optional[List[str]] = None,
    expected_type: Optional[str] = None,
    edf_filters: Optional[Dict] = None,
):
    _require_bucket()
    if not pdf_keys:
        raise RuntimeError("Missing pdf_keys")

    if min_budget_m is None:
        min_budget_m = DEFAULT_MIN_BUDGET_M

    original_names = original_names or []
    edf_filters = edf_filters or {}

    all_rows: List[Dict] = []
    detected_type: Optional[str] = None

    for idx, key in enumerate(pdf_keys):
        local_pdf = f"/tmp/{uuid.uuid4()}.pdf"
        s3.download_file(BUCKET, key, local_pdf)

        text = extract_text(local_pdf)
        doc_type = detect_document_family(text)
        file_label = original_names[idx] if idx < len(original_names) else key

        if doc_type == "unknown":
            raise RuntimeError(f"Unable to detect document family from PDF text ({file_label}).")

        if expected_type and doc_type != expected_type:
            raise RuntimeError(
                f"Uploaded document looks like {doc_type.upper()} but active tab is {expected_type.upper()} ({file_label})."
            )

        if detected_type and doc_type != detected_type:
            raise RuntimeError("Cannot mix Horizon and EDF PDFs in the same request.")
        detected_type = doc_type

        parsed_rows = parse_calls(text) if doc_type == DOC_HORIZON else parse_edf(text)
        for r in parsed_rows:
            r["source_pdf"] = file_label
        all_rows.extend(parsed_rows)

    if not detected_type:
        raise RuntimeError("No documents processed.")

    # Shared temp Excel path
    local_xlsx = f"/tmp/{uuid.uuid4()}.xlsx"

    if detected_type == DOC_EDF:
        call_family = (edf_filters.get("call_family") or "").strip()
        budget_min = _coerce_float(edf_filters.get("budget_min_m"))
        budget_max = _coerce_float(edf_filters.get("budget_max_m"))
        step_filter = _coerce_bool(edf_filters.get("step"))

        # Ensure derived call_family exists
        call_types_meta = {}
        for r in all_rows:
            r["call_family"] = r.get("call_family") if r.get("call_family") in EDF_CALL_FAMILY_LABELS else _edf_call_family_from_id(r.get("call_id"))
            is_large_scale = r.get("is_large_scale")
            if is_large_scale is None:
                is_large_scale = False
            r["is_large_scale"] = bool(is_large_scale or _edf_is_large_scale(r))
            r["call_family_display"] = _edf_call_family_label(r.get("call_family")) or r.get("call_family")
            r["scale"] = _edf_scale_label(r["is_large_scale"])
            r["budget_per_project_min_eur_m"] = r.get("budget_per_project_min_eur_m") or r.get("indicative_budget_eur_m")
            r["call_type"] = r.get("type_of_action") or r.get("call_family_display")
            r["funding_percentage"] = _funding_percentage(r, DOC_EDF)
            if r["call_type"] and r["call_type"] not in call_types_meta:
                call_types_meta[r["call_type"]] = r["funding_percentage"]

        rows = [r for r in all_rows if r.get("record_level") == "TOPIC"]
        rows = filter_edf_rows(
            rows,
            call_family=call_family,
            budget_min_m=budget_min,
            budget_max_m=budget_max,
            step=step_filter,
        )

        # Apply shared filters (call types, budget slider, opening/deadline)
        rows = filter_rows(
            rows,
            call_types=call_types,
            min_budget_m=min_budget_m,
            opening_filter=opening_filter,
            deadline_filter=deadline_filter,
            doc_type=detected_type,
        )

        write_xlsx(rows + [r for r in all_rows if r.get("record_level") == "CALL"], local_xlsx, DOC_EDF)

        safe_base = _safe_base_name(original_names[0] if original_names else pdf_keys[0])
        if len(pdf_keys) > 1:
            safe_base = f"{safe_base}-combined"
        out_key = f"outputs/{uuid.uuid4()}/{safe_base}.xlsx"
        s3.upload_file(local_xlsx, BUCKET, out_key)

        display_rows = []
        for r in rows:
            call_type = _row_call_type(r, DOC_EDF)
            funding = _funding_percentage(r, DOC_EDF)
            if call_type and call_type not in call_types_meta:
                call_types_meta[call_type] = funding

            display_rows.append(
                {
                    "record_level": r.get("record_level"),
                    "call_id": r.get("call_id"),
                    "topic_id": r.get("topic_id"),
                    "topic_id": r.get("topic_id"),
                    "topic_url": _topic_url(r.get("topic_id")),
                    "topic_title": r.get("topic_title"),
                    "title": r.get("title"),
                    "section_no": r.get("section_no"),
                    "budget_per_project_min_eur_m": r.get("budget_per_project_min_eur_m"),
                    "opening_date": r.get("opening_date"),
                    "deadline_date": r.get("deadline_date"),
                    "call_type": call_type,
                    "funding_percentage": funding,
                }
            )

        return {
            "status": "ok",
            "excel_key": out_key,
            "rows": display_rows,
            "rows_count": len(rows),
            "doc_type": detected_type,
            "call_types": [{"name": k, "funding_percentage": v} for k, v in call_types_meta.items()],
        }

    # Horizon flow
    call_types_meta = {}
    for r in all_rows:
        derived_budget = _compute_budget_per_project_m(r)
        r["budget_per_project_min_eur_m"] = derived_budget
        r["budget_per_project_m"] = derived_budget
        r["funding_percentage"] = _funding_percentage(r, DOC_HORIZON)
        call_type = _row_call_type(r, DOC_HORIZON)
        if call_type and call_type not in call_types_meta:
            call_types_meta[call_type] = r["funding_percentage"]

    rows = filter_rows(
        all_rows,
        call_types=call_types,
        min_budget_m=min_budget_m,
        opening_filter=opening_filter,
        deadline_filter=deadline_filter,
        doc_type=detected_type,
    )

    # --- OpenAI descriptions (optional) ---
    desc_cache: dict = {}
    if OPENAI_API_KEY:
        n = 0
        for r in rows:
            if n >= OPENAI_MAX_TOPICS:
                break

            if context is not None:
                remaining_ms = context.get_remaining_time_in_millis()
                if remaining_ms is not None and remaining_ms < 8000:
                    print(f"Stopping OpenAI: low remaining time {remaining_ms}ms")
                    break

            body = (r.get("topic_body") or "").strip()
            if not body:
                continue

            cur = (r.get("topic_description") or "").strip()
            if len(cur) >= 80:
                continue

            desc = _openai_topic_description(
                topic_id=r.get("topic_id") or "",
                topic_title=r.get("topic_title") or "",
                body_text=body,
                cache=desc_cache,
            )
            if desc:
                r["topic_description"] = desc
                n += 1
            else:
                r["topic_description"] = ""

    for r in rows:
        if r.get("topic_description") is None:
            r["topic_description"] = ""

        if r.get("budget_per_project_min_eur_m") is None:
            derived_budget = _compute_budget_per_project_m(r)
            r["budget_per_project_min_eur_m"] = derived_budget
            r["budget_per_project_m"] = derived_budget

        r.pop("topic_body", None)

    write_xlsx(rows, local_xlsx, DOC_HORIZON)

    safe_base = _safe_base_name(original_names[0] if original_names else pdf_keys[0])
    if len(pdf_keys) > 1:
        safe_base = f"{safe_base}-combined"
    out_key = f"outputs/{uuid.uuid4()}/{safe_base}.xlsx"
    s3.upload_file(local_xlsx, BUCKET, out_key)

    display_rows = []
    for r in rows:
        call_type = _row_call_type(r, DOC_HORIZON)
        funding = _funding_percentage(r, DOC_HORIZON)
        if call_type and call_type not in call_types_meta:
            call_types_meta[call_type] = funding

        display_rows.append(
            {
                "topic_id": r.get("topic_id"),
                "topic_url": _topic_url(r.get("topic_id")),
                "topic_title": r.get("topic_title") or "",
                "budget_per_project_min_eur_m": r.get("budget_per_project_min_eur_m"),
                "opening_date": r.get("opening_date"),
                "deadline_date": r.get("deadline_date"),
                "call_type": call_type,
                "funding_percentage": funding,
            }
        )

    return {
        "status": "ok",
        "excel_key": out_key,
        "rows": display_rows,
        "rows_count": len(rows),
        "doc_type": detected_type,
        "call_types": [{"name": k, "funding_percentage": v} for k, v in call_types_meta.items()],
    }
def extract_text(pdf_path: str) -> str:
    """
    Extract text with explicit page markers so parser_horizon can set 'page'.
    """
    reader = PdfReader(pdf_path)
    chunks = []
    for idx, p in enumerate(reader.pages, start=1):
        chunks.append(f"\n<<<PAGE {idx}>>>\n")
        chunks.append(p.extract_text() or "")
    return "\n".join(chunks)


def _compute_budget_per_project_m(row):
    vals = []
    for key in (
        "budget_per_project_min_eur_m",
        "budget_per_project_max_eur_m",
        "budget_per_project_m",
    ):
        v = row.get(key)
        if isinstance(v, (int, float)):
            vals.append(float(v))
    if vals:
        return min(vals)
    return None


def _matches_prefix(value: str, prefix: str) -> bool:
    pref = (prefix or "").strip()
    if not pref:
        return True
    if value is None:
        return False
    return str(value).strip().lower().startswith(pref.lower())


def detect_document_family(text: str) -> str:
    low = (text or "").lower()
    horizon_signal = (
        "horizon europe" in low
        or "work programme" in low
        or re.search(r"\bhorizon-[a-z0-9]+-\d{4}-", text, flags=re.IGNORECASE)
    )
    edf_matches = re.findall(r"\bedf-\d{4}-[a-z]{2,}", text, flags=re.IGNORECASE)
    edf_keyword = "european defence fund" in low
    if horizon_signal:
        return DOC_HORIZON
    strong_edf = (edf_keyword and bool(edf_matches)) or len(edf_matches) >= 3
    if strong_edf:
        return DOC_EDF
    return "unknown"


def _write_horizon_xlsx(rows, xlsx_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "calls"

    headers = [
        "cluster",
        "stage",
        "call_round",
        "page",
        "call_id",
        "topic_id",
        "topic_title",
        "topic_description",
        "action_type",
        "funding_percentage",
        "trl",
        "budget_eur_m",
        "budget_per_project_min_eur_m",
        "budget_per_project_max_eur_m",
        "projects",
        "opening_date",
        "deadline_date",
    ]
    ws.append(headers)

    for r in rows:
        row_values = [r.get(h) for h in headers]
        ws.append(row_values)

    # apply hyperlink to topic_id column (6th col)
    for idx, r in enumerate(rows, start=2):  # header is row 1
        cell = ws.cell(row=idx, column=6)
        url = _topic_url(r.get("topic_id"))
        if url:
            cell.hyperlink = url
            cell.style = "Hyperlink"

    wb.save(xlsx_path)


def _write_edf_xlsx(rows, xlsx_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "edf"

    headers = [
        "record_level",
        "call_id",
        "call_family",
        "call_family_display",
        "topic_id",
        "title",
        "topic_title",
        "section_no",
        "type_of_action",
        "funding_percentage",
        "budget_per_project_min_eur_m",
        "indicative_budget_eur_m",
        "call_indicative_budget_eur_m",
        "number_of_actions",
        "step",
        "scale",
        "is_large_scale",
        "topic_description_verbatim",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([r.get(h) for h in headers])

    # Wrap long verbatim descriptions
    wrap_align = Alignment(wrap_text=True, vertical="top")
    desc_col_idx = headers.index("topic_description_verbatim") + 1
    desc_col_letter = ws.cell(row=1, column=desc_col_idx).column_letter
    ws.column_dimensions[desc_col_letter].width = 100

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=desc_col_idx).alignment = wrap_align

    wb.save(xlsx_path)


def write_xlsx(rows, xlsx_path: str, doc_type: str):
    if doc_type == DOC_EDF:
        _write_edf_xlsx(rows, xlsx_path)
    else:
        _write_horizon_xlsx(rows, xlsx_path)


def _parse_date(s: str):
    """
    Parse YYYY, YYYY-MM, or YYYY-MM-DD into a date object.
    Also supports textual dates like "23 Sep 2026" or "23 September 2026".
    Returns None when parsing fails.
    """
    txt = (s or "").strip()
    if not txt:
        return None

    # Remove trailing punctuation that often appears in PDF extracts
    txt = txt.rstrip(".,;")

    m_full = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", txt)
    if m_full:
        y, mo, d = int(m_full.group(1)), int(m_full.group(2)), int(m_full.group(3))
        try:
            return date(y, mo, d)
        except ValueError:
            return None

    m_month = re.match(r"^(\d{4})-(\d{2})$", txt)
    if m_month:
        y, mo = int(m_month.group(1)), int(m_month.group(2))
        try:
            return date(y, mo, calendar.monthrange(y, mo)[1])
        except ValueError:
            return None

    m_year = re.match(r"^(\d{4})$", txt)
    if m_year:
        y = int(m_year.group(1))
        try:
            return date(y, 12, 31)
        except ValueError:
            return None

    m_day_name = re.match(r"^(\d{1,2})\s+([A-Za-z]{3,})\.?,?\s+(\d{4})$", txt)
    if m_day_name:
        day = int(m_day_name.group(1))
        mon_raw = m_day_name.group(2).strip().lower().rstrip(".")
        year = int(m_day_name.group(3))
        months = {
            "jan": 1, "january": 1,
            "feb": 2, "february": 2,
            "mar": 3, "march": 3,
            "apr": 4, "april": 4,
            "may": 5,
            "jun": 6, "june": 6,
            "jul": 7, "july": 7,
            "aug": 8, "august": 8,
            "sep": 9, "sept": 9, "september": 9,
            "oct": 10, "october": 10,
            "nov": 11, "november": 11,
            "dec": 12, "december": 12,
            # Italian month names (PDFs sometimes localized)
            "gen": 1, "gennaio": 1,
            "febbraio": 2,
            "marzo": 3,
            "aprile": 4,
            "maggio": 5,
            "giugno": 6,
            "luglio": 7,
            "agosto": 8,
            "settembre": 9,
            "ottobre": 10,
            "novembre": 11,
            "dicembre": 12,
        }
        mo = months.get(mon_raw)
        if mo:
            try:
                return date(year, mo, day)
            except ValueError:
                return None

    m_day_slash = re.match(r"^(\d{1,2})[./](\d{1,2})[./](\d{4})$", txt)
    if m_day_slash:
        d, mo, y = int(m_day_slash.group(1)), int(m_day_slash.group(2)), int(m_day_slash.group(3))
        try:
            return date(y, mo, d)
        except ValueError:
            return None

    return None


def _parse_filter_range(filter_value: str):
    """
    Convert user filter into an (start, end) tuple.
    - YYYY -> (None, 31 Dec YYYY)
    - YYYY-Qx -> (None, end_of_quarter)
    - YYYY-MM -> (None, end_of_month)
    - YYYY-MM-DD -> (None, that day)
    Returns None if the filter is empty or invalid.
    """
    txt = (filter_value or "").strip()
    if not txt:
        return None

    m_year = re.match(r"^(\d{4})$", txt)
    if m_year:
        y = int(m_year.group(1))
        try:
            return (None, date(y, 12, 31))
        except ValueError:
            return None

    m_quarter = re.match(r"^(\d{4})-Q([1-4])$", txt, flags=re.IGNORECASE)
    if m_quarter:
        y = int(m_quarter.group(1))
        q = int(m_quarter.group(2))
        month_end = q * 3
        try:
            end_day = calendar.monthrange(y, month_end)[1]
            return (None, date(y, month_end, end_day))
        except ValueError:
            return None

    m_month = re.match(r"^(\d{4})-(\d{2})$", txt)
    if m_month:
        y, mo = int(m_month.group(1)), int(m_month.group(2))
        try:
            end_day = calendar.monthrange(y, mo)[1]
            return (None, date(y, mo, end_day))
        except ValueError:
            return None

    m_day = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", txt)
    if m_day:
        y, mo, d = int(m_day.group(1)), int(m_day.group(2)), int(m_day.group(3))
        try:
            return (None, date(y, mo, d))
        except ValueError:
            return None

    return None


def _date_filter_match(value: str, filter_value: str) -> bool:
    """
    Match dates using inclusive upper-bound logic:
    - If filter is a valid date/period, include rows with dates <= end_of_period.
    - Otherwise, fallback to prefix match to avoid breaking existing inputs.
    """
    rng = _parse_filter_range(filter_value)
    if not rng:
        return _matches_prefix(value, filter_value)

    _start, end = rng
    row_date = _parse_date(value)
    if not row_date:
        return False

    if end and row_date > end:
        return False
    return True


def filter_rows(
    rows,
    call_types=None,
    min_budget_m: float = None,
    opening_filter: str = "",
    deadline_filter: str = "",
    doc_type: str = DOC_HORIZON,
):
    allowed = None
    if call_types:
        allowed = {str(t).strip().lower() for t in call_types if str(t).strip()}
        if not allowed:
            allowed = None

    filtered = []
    for r in rows:
        if allowed is not None:
            cur = (_row_call_type(r, doc_type) or "").strip().lower()
            if cur not in allowed:
                continue

        if min_budget_m is not None:
            budget_val = _row_min_budget(r)
            if budget_val is None:
                budget_val = 0.0
            if budget_val < min_budget_m:
                continue

        if not _date_filter_match(r.get("opening_date"), opening_filter):
            continue

        if not _date_filter_match(r.get("deadline_date"), deadline_filter):
            continue

        filtered.append(r)

    return filtered


def filter_edf_rows(
    rows,
    call_family: str = "",
    budget_min_m: float = None,
    budget_max_m: float = None,
    step: Optional[bool] = None,
):
    fam = (call_family or "").strip().lower()

    filtered = []
    for r in rows:
        if fam:
            cur = (r.get("call_family") or "").lower()
            if not cur.startswith(fam):
                continue

        budget_val = r.get("indicative_budget_eur_m")
        if budget_min_m is not None or budget_max_m is not None:
            if not isinstance(budget_val, (int, float)):
                continue
            if budget_min_m is not None and budget_val < budget_min_m:
                continue
            if budget_max_m is not None and budget_val > budget_max_m:
                continue

        if step is not None:
            cur_step = r.get("step")
            if cur_step is None:
                continue
            if bool(cur_step) != bool(step):
                continue

        filtered.append(r)
    return filtered


# --- OpenAI helpers (Responses API) ---
def _extract_output_text(resp_json: dict) -> str:
    # Some responses include output_text directly
    txt = (resp_json.get("output_text") or "").strip()
    if txt:
        return txt

    # Otherwise parse output array
    out = resp_json.get("output") or []
    parts = []
    for item in out:
        content = item.get("content") or []
        for c in content:
            # depending on the exact shape, you may see output_text or text
            if c.get("type") == "output_text" and c.get("text"):
                parts.append(c["text"])
            elif c.get("type") == "text" and c.get("text"):
                parts.append(c["text"])
    return "\n".join(p.strip() for p in parts if p and p.strip()).strip()


def _openai_topic_description(topic_id: str, topic_title: str, body_text: str, cache: dict) -> str:
    """
    Return a short English summary (max 2 sentences) using ONLY the provided text.
    """
    if not OPENAI_API_KEY:
        return ""

    clean_body = (body_text or "").strip()
    if not clean_body:
        return ""

    if len(clean_body) > OPENAI_BODY_MAX_CHARS:
        clean_body = clean_body[:OPENAI_BODY_MAX_CHARS]

    if clean_body in cache:
        return cache[clean_body]

    instructions = (
        "Summarize only what is present in the provided text.\n"
        "Do not add requirements, dates, numbers, or claims not present.\n"
        "Return 2 short sentences maximum in English using simple language (under 240 characters)."
    )

    user_input = (
        f"Topic ID: {topic_id}\n"
        f"Title: {topic_title}\n\n"
        "Text from PDF:\n"
        f"{clean_body}"
    )

    payload = {
        "model": OPENAI_MODEL,
        "instructions": instructions,
        "input": user_input,
        "store": False,
    }

    req = urllib.request.Request(
        "https://api.openai.com/v1/responses",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            data = json.loads(r.read().decode("utf-8"))
            summary = _extract_output_text(data).strip()
            if summary:
                sentences = re.split(r"(?<=[.!?])\s+", summary)
                summary = " ".join([p for p in sentences if p][:2]).strip()
            if len(summary) > 240:
                summary = summary[:240].rstrip()
            cache[clean_body] = summary
            return summary
    except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError):
        cache[clean_body] = ""
        return ""
    except Exception:
        cache[clean_body] = ""
        return ""


# HTML served from ui.html loaded at startup

def _resp(status_code: int, body: str, content_type: str = "application/json"):
    return {
        "statusCode": status_code,
        "headers": {
            "content-type": content_type,
            "access-control-allow-origin": "*",
            "access-control-allow-methods": "GET,POST,OPTIONS",
            "access-control-allow-headers": "content-type",
        },
        "body": body,
    }


def _json(obj):
    return json.dumps(obj, ensure_ascii=False)


def _coerce_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _coerce_bool(val):
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        low = val.strip().lower()
        if low in {"true", "yes", "1", "y", "on"}:
            return True
        if low in {"false", "no", "0", "off"}:
            return False
    return None


def handler(event, context):
    try:
        # Supporta invocazioni "dirette" (CLI) e HTTP (Lambda URL)
        if "requestContext" not in event:
            pdf_keys = event.get("pdf_keys") or []
            if not pdf_keys and event.get("pdf_key"):
                pdf_keys = [event["pdf_key"]]
            call_types = event.get("call_types") or event.get("action_types")
            min_budget_m = _coerce_float(event.get("min_budget_m"))
            if min_budget_m is None:
                min_budget_m = DEFAULT_MIN_BUDGET_M
            original_names = event.get("original_names") or []
            opening_filter = event.get("opening_filter") or ""
            deadline_filter = event.get("deadline_filter") or ""
            expected_type = event.get("expected_type") or event.get("doc_family")
            edf_filters = event.get("edf_filters") or {}
            return _process_pdf_keys(
                pdf_keys,
                context=context,
                call_types=call_types,
                min_budget_m=min_budget_m,
                opening_filter=opening_filter,
                deadline_filter=deadline_filter,
                original_names=original_names,
                expected_type=expected_type,
                edf_filters=edf_filters,
            )

        method = event.get("requestContext", {}).get("http", {}).get("method", "GET")
        path = event.get("rawPath", "/")

        if method == "OPTIONS":
            return _resp(200, "")

        # Serve static assets packaged with the Lambda
        if method == "GET" and path.startswith("/assets/"):
            return _serve_asset(path)

        if method == "GET" and path == "/":
            return _resp(200, HTML, content_type="text/html; charset=utf-8")

        if method == "GET" and path == "/presign":
            _require_bucket()
            params = event.get("queryStringParameters") or {}
            try:
                count = int(params.get("count") or "1")
            except ValueError:
                count = 1
            count = max(1, min(6, count))

            uploads = []
            for _ in range(count):
                pdf_key = f"uploads/{uuid.uuid4()}.pdf"
                upload_url = s3.generate_presigned_url(
                    "put_object",
                    Params={"Bucket": BUCKET, "Key": pdf_key},
                    ExpiresIn=900,
                )
                uploads.append({"upload_url": upload_url, "pdf_key": pdf_key})

            payload = {"uploads": uploads}
            if uploads:
                payload["upload_url"] = uploads[0]["upload_url"]
                payload["pdf_key"] = uploads[0]["pdf_key"]
            return _resp(200, _json(payload))

        if method == "POST" and path == "/process":
            _require_bucket()
            body = event.get("body") or "{}"
            data = json.loads(body)
            pdf_keys = data.get("pdf_keys") or []
            if not pdf_keys and data.get("pdf_key"):
                pdf_keys = [data["pdf_key"]]
            call_types = data.get("call_types") or data.get("action_types")
            min_budget_m = _coerce_float(data.get("min_budget_m"))
            if min_budget_m is None:
                min_budget_m = DEFAULT_MIN_BUDGET_M
            original_names = data.get("original_names") or []
            opening_filter = data.get("opening_filter") or ""
            deadline_filter = data.get("deadline_filter") or ""
            expected_type = data.get("expected_type") or data.get("doc_family")
            edf_filters = data.get("edf_filters") or {}
            result = _process_pdf_keys(
                pdf_keys,
                context=context,
                call_types=call_types,
                min_budget_m=min_budget_m,
                opening_filter=opening_filter,
                deadline_filter=deadline_filter,
                original_names=original_names,
                expected_type=expected_type,
                edf_filters=edf_filters,
            )
            return _resp(200, _json(result))

        if method == "POST" and path == "/download":
            _require_bucket()
            body = event.get("body") or "{}"
            data = json.loads(body)
            excel_key = data["excel_key"]
            download_url = s3.generate_presigned_url(
                "get_object",
                Params={"Bucket": BUCKET, "Key": excel_key},
                ExpiresIn=900,
            )
            return _resp(200, _json({"download_url": download_url}))

        return _resp(404, _json({"error": "not_found", "path": path, "method": method}))

    except Exception as e:
        # log completo su CloudWatch, ma rispondiamo con messaggio leggibile al browser
        print("ERROR:", repr(e))
        print(traceback.format_exc())

        return _resp(
          500,
          _json({
          "error": "internal",
          "message": str(e),
          "type": e.__class__.__name__,
          "request_id": getattr(context, "aws_request_id", None),
          }),
        )
